import base64
import shutil
from datetime import datetime
import os
import sys
import time
from tkinter import messagebox
import tkinter as tk
import ctypes
from io import StringIO, BytesIO
from PIL import ImageTk, Image
import openpyxl
import os
from pathlib import Path

from Codes.MailSmtp import MailSmtp
from Codes.pic2str import sendmail,maillogo


class MailCustomer:
    def __init__(self):
        pass

    def Gui(self):

        global label_status
        global textbox_schemedetails

        import tkinter as tk
        from PIL import Image, ImageTk
        import base64
        from io import BytesIO

        ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)

        window = tk.Tk()

        window.configure(background='#00AAE4')
        window.title("MailCustomer_CNDetails")

        window.columnconfigure(0, weight=1, minsize=75)
        window.rowconfigure(0, weight=1, minsize=50)
        window.columnconfigure(1, weight=1, minsize=75)
        window.rowconfigure(1, weight=1, minsize=50)
        window.columnconfigure(2, weight=1, minsize=75)
        window.rowconfigure(2, weight=1, minsize=50)

        w = 350
        h = 400
        x = 900
        y = 0

        window.geometry('%dx%d+%d+%d' % (w, h, x, y))

        frame_image = tk.Frame(
            master=window,
            relief=tk.RAISED,
            background='#00AAE4',
        )

        byte_data = base64.b64decode(maillogo)
        image_data = BytesIO(byte_data)
        image = Image.open(image_data)
        resize_image = image.resize((90, 70))
        img = ImageTk.PhotoImage(resize_image)
        label2 = tk.Label(master=frame_image, image=img)
        label2.pack()

        frame_schemedetails = tk.Frame(
            master=window,
            relief=tk.RAISED,
            background='#00AAE4',
        )

        label_schemedetails = tk.Label(
            master=frame_schemedetails,
            text="Scheme details:",
            fg="white",
            bg="#00AAE4",
            font=('Arial', 13)
        )
        label_schemedetails.pack()

        textbox_schemedetails = tk.Text(
            master=frame_schemedetails,
            width=30,
            height=1,
            font=('Arial', 11)
        )
        textbox_schemedetails.pack()

        frame_button = tk.Frame(
            master=window,
            relief=tk.RAISED,
            background='#00AAE4',
        )

        button = tk.Button(
            master=frame_button,
            width=15,
            height=2,
            text="SendMail",
            command=self.sendmail
        )
        button.pack()

        frame_status = tk.Frame(
            master=window,
            relief=tk.RAISED,
            background='#00AAE4',
        )

        label_status = tk.Label(master=frame_status, text="", fg="white", bg="#00AAE4", font=('Arial', 13))
        label_status.pack()

        frame_image.grid(row=0, column=0)
        frame_schemedetails.grid(row=1, column=1)
        frame_button.grid(row=2, column=1)
        frame_status.grid(row=3, column=1)
        window.mainloop()

    def sendmail(self):

        cwd_os = os.getcwd()
        communication = str(Path.cwd()) + str("\\Communication")
        inputwkpath = str(Path.cwd()) + str("\\Email_Communication.xlsm")
        scheme_details = textbox_schemedetails.get("1.0", tk.END)

        # Load the workbook
        workbook = openpyxl.load_workbook(inputwkpath, keep_vba=True)
        first_sheet_name = workbook.sheetnames[1]
        sheet = workbook[first_sheet_name]

        # Iterate through the rows
        for i in range(2,10000):

            if str(sheet.cell(row=i,column=1).value)!="None":

                customercode = (str(sheet.cell(row=i,column=1).value))
                customername = (str(sheet.cell(row=i,column=2).value))
                cnrefno = (str(sheet.cell(row=i,column=3).value))
                partnermail = (str(sheet.cell(row=i,column=4).value))
                ccdmail = (str(sheet.cell(row=i,column=5).value))

                smtpmail = MailSmtp(customercode,customername,cnrefno,partnermail,ccdmail,communication,scheme_details)
                smtpmail.sendmailtocustomers()





